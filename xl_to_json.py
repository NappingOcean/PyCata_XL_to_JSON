'''

구현 목표
1. xlsx 로부터 데이터를 가져와서 JSON 파일을 만든다.

2. JSON 파일로부터 형식만을 취하여 템플릿 xlsx 파일을 만든다.

3. 객체와 배열을 구분하여 데이터를 넣는다.
    - {}로서 표현되는 객체와 []로서 표현되는 배열.
    - 그럼 엑셀은 어떤 식으로 작성해둬야 하는가? 이는 나중에 마크다운 파일에 적어놓을 것.

4. JSON 파일이 들어있는 폴더를 연다.   

5. 이상의 사항들을 이벤트 루프 속에서 처리하는 창을 따로 띄우도록 한다.

'''

#import zone

import openpyxl as opyxl
import json as jo
import os

#import zone



class Data_Converter:
    def __init__(self):
        pass
        
    # 1행에 있는 키 중에 복잡한 키를 분리하여 리스트화.
    def key_separator(self, raw_keys) -> list:
        # 받은 keys를 / 별로 짤라서 리스트로 저장한다.
        key_serial = raw_keys.split('/')
        return key_serial
    
    # 본격적으로 딕셔너리 구축            
    def dix_builder(self, keys, vals) -> dict:
        
        #k_map = jagged list
        k_map = list(map(self.key_separator, keys))
        
        k_last_list = list(map(lambda ks:ks[-1], k_map))
        dix_last = dict(zip(k_last_list, vals))
        
        bld_dix = {}

        for k_serial in k_map:
            tmp_idx = 0
            k_idx = 0
            for k in reversed(k_serial):
                k_idx-=1
                k_cut = k.removesuffix(':list')
                if k == k_serial[-1]:
                    tmp_dix = {}
                    if k.endswith(':list') and not isinstance(dix_last[k], list):
                        tmp_dix = {k_cut:[dix_last[k]]}
                    else:
                        tmp_dix = {k_cut:dix_last[k]}
                else: # k가 맨 끝에 있는 k 가 아니다.
                    
                    if not k in [k_ser[k_idx] for k_ser in k_map[:k_map.index(k_serial)] if len(k_ser)+k_idx >= 0 and k_ser[:k_idx] == k_serial[:k_idx]]:
                        if k.endswith(':list'):
                            tmp_dix = {k_cut:[tmp_dix]}
                        else:
                            tmp_dix = {k:tmp_dix}
                    else:
                        tmp_idx += 1

            # end of for-loop: k in k_serial
            tmp_adr = bld_dix

            if tmp_idx > 0: # tmp_dix 의 디렉토리와 겹치는 값이 하나 이상!
                for k2 in k_serial[:tmp_idx]:
                    k2_cut = k2.removesuffix(':list')
                    # 기존 값의 어디에 붙여야 하는지 탐색.
                    if not isinstance(tmp_adr,list):
                        tmp_adr = tmp_adr[k2_cut]
                    else:
                        for dix in tmp_adr:
                            if k2_cut in dix:
                                tmp_adr = dix

                    if k2 == k_serial[tmp_idx-1]:
                        if isinstance(tmp_adr,list):
                            tmp_adr.append(tmp_dix)
                        else:
                            if not isinstance(tmp_adr,list):
                                tmp_adr.update(tmp_dix)
                            else:
                                tmp_adr.append(tmp_dix)
            else:
                tmp_adr.update(tmp_dix)

        # end of for-loop:k_serial in k_map

        # Delete the key with an empty value inside
        for key in list(bld_dix.keys()):
            if bld_dix[key] == None:
                del(bld_dix[key])
        
        return bld_dix

    def json_to_templete(self):
            pass
    
 
class File_Loader_Opyxl:
    def __init__(self, filename:str):
        self.fn = filename

    def read_file(self):
        opyxl_file = opyxl.load_workbook(self.fn, data_only=True)
        
        return opyxl_file
    
    def raw_keys_from_file(self, opyxl_file:opyxl.Workbook) -> dict: 
        
        keys_dict = {}

        for s_name in opyxl_file.sheetnames:
            ws = opyxl_file[s_name]
            list_from_ws = [ws.cell(row = 1, column= i+1).value for i in range(ws.max_column)]
            # for idx in range(len(list_from_ws)):
            #     if list_from_ws[idx].startswith('Unnamed: '):
            #         list_from_ws[idx] = None
                    
            keys_dict[s_name] = list_from_ws

        return keys_dict
    
    def keys_pro(self, opyxl_file:opyxl.Workbook):

        keys_dict = {}

        r_keys = self.raw_keys_from_file(opyxl_file)

        for name in opyxl_file.sheetnames:
            keys_dict[name] = list(filter(None, r_keys[name]))

        return keys_dict


    def raw_vals_from_file(self, opyxl_file:opyxl.Workbook) -> dict: 
        
        raw_vals_dict = {}

        for s_name in opyxl_file.sheetnames:
            ws = opyxl_file[s_name]
            
            list_from_ws = [[ws.cell(row=y+1, column=x+1).value for x in range(ws.max_column)] for y in range(ws.max_row)]
            # for lst in list_from_ws:
            #     for idx in range(len(lst)):
            #         if pd.isna(lst[idx]):
            #             lst[idx] = None
            raw_vals_dict[s_name] = list_from_ws[1:]

        return raw_vals_dict
    
    # 주어진 row 이하의 col 에서 값은 얼마나 있는가?
    def amount_checker(self, lst:list, row, col) -> int:
        ser_num = 1
        for chk_row in range(row + 1, len(lst)):
            if lst[chk_row][col] != None and lst[chk_row][0] == None:
                #col 값이 비어있지 않으면서 id 는 비어있는 경우.
                ser_num += 1
            else:
                break
        return ser_num

    def vals_pro(self, opyxl_file:opyxl.Workbook):

        vals_dict = {}

        keys = self.raw_keys_from_file(opyxl_file)
        r_vals = self.raw_vals_from_file(opyxl_file)

        for s_name in opyxl_file.sheetnames:
            key_list = keys[s_name]
            val_list_list = r_vals[s_name]
            
            val_dict_in_sht = {}
            
            for row in range(len(val_list_list)):
                val_list = val_list_list[row]
                val_id = val_list[0]
                val_dict_in_row = {}
                tmp_val = []
                for col in range(len(key_list)):
                    key_now = key_list[col]
                    val_now = val_list[col]
                    if val_id:
                        if key_now:
                            val_dict_in_row[key_now] = val_now
                            key_pre = key_now
                        elif val_now != None:
                            tmp_val = val_dict_in_row[list(val_dict_in_row.keys())[-1]]
                            if isinstance(tmp_val, list):
                                tmp_val.append(val_now)
                            else:
                                tmp_val = [tmp_val, val_now]
                                val_dict_in_row.update({key_pre:tmp_val})
                        else: #id 만 있고 키도 값도 없는 자리.
                            # 거지같은 vitamins_absorb_multi 같으니.
                            val_up_pre = val_dict_in_row[key_pre]
                            no_id_val = [[val_list_list[tmp_row][col]] for tmp_row in range(row+1, row + self.amount_checker(val_list_list, row, col))] # [[w#],[x#]]
                            if not isinstance(val_up_pre, list):
                                val_up_pre = [val_up_pre, no_id_val] # [v]
                                val_dict_in_row[key_pre] = [val_up_pre]
                                # [ v,[[w0],[x0]] ]
                            else:
                                # for i in range(len(val_up_pre[-1])):
                                #     val_up_pre[-1][i].extend(no_id_val[i])
                                val_dict_in_row[key_pre][0][-1] =[[val_up_pre[0][-1][i][0],no_id_val[i][0]] for i in range(len(val_up_pre[0][-1]))]
                                    # [ v,[[w0, w1],[x0,x1]] ]
                        id_pre = val_id
                    
                    else:   # 그 자리 id 없음
                        if val_now != None: # 그 자리 값 있음
                            if key_now: # 그 자리 키 있음
                                upper_val = val_dict_in_sht[id_pre][key_now]
                                if isinstance(upper_val, list):
                                    # 다음 키가 비어있지 않으면.
                                    if key_now == key_list[-1] or key_list[key_list.index(key_now)+1]:
                                        upper_val.append(val_now)
                                    elif isinstance(upper_val[0], list):
                                        val_now = [val_list[col+i] for i in range(len(upper_val[0]))]
                                        upper_val.append(val_now)
                                    else: # 다음키 빔. 첫요소 리스트 아님.
                                        val_now = [val_list[col+i] for i in range(len(upper_val))]
                                        val_dict_in_sht[id_pre][key_now] = [upper_val]
                                        val_dict_in_sht[id_pre][key_now].append(val_now)
                                else: # upper_val이 리스트가 아니면.
                                    val_dict_in_sht[id_pre][key_now] = [upper_val, val_now]
                                key_pre = key_now
                # end of for-loop:col
                val_dict_in_sht.update({val_id:val_dict_in_row})
            # end of for-loop:row
            vals_dict[s_name] = list(filter(None, [val_dict_in_sht[key] for key in val_dict_in_sht.keys()])) # val_dict_in_sht:dict 중에서 비어있는 딕션을 없앤다.
            vals_dict[s_name] = [list(vals_dict[s_name][i].values()) for i in range(len(vals_dict[s_name]))] #vals_dict 를 list를 value로 가진 dict로 재정의한다.
        # end of for-loop:s_name
        return vals_dict
    
class File_Open_Path:
    def __init__(self, xl_path:str = './Excel/', jo_path:str = './JSON/'):
        self.xp = xl_path
        self.jp = jo_path

    def jo_writer(self, ls_data:list, jo_name:str)-> None:
        name = jo_name + ".json"
        with open(self.jp + name, 'w', encoding="utf-8") as file:
            jo.dump(ls_data, file, indent=3)
       
    def xl_reader(self) -> str:
        return self.xp
    
class Xl_to_Json:
    def __init__(self, filename_xl, xl_path:str = './Excel/', jo_path:str = './JSON/'):
        self.fn_xl = filename_xl
        self.xp = xl_path
        self.jp = jo_path

    def conv(self):


        main_xl = File_Loader_Opyxl(self.fn_xl)

        main_wb = main_xl.read_file()
        m_k_fr_xl = main_xl.keys_pro(main_wb)
        m_v_fr_xl = main_xl.vals_pro(main_wb)

        m_xl_conv = Data_Converter()

        for name in list(m_v_fr_xl.keys()):
            data_list = []
            for i in range(len(m_v_fr_xl[name])):
                data = {'type':name}
                data.update(m_xl_conv.dix_builder(m_k_fr_xl[name], m_v_fr_xl[name][i]))
                data_list.append(data)
            # end of for-loop:i
            xl_jo_dir = self.jp + self.fn_xl[:-5]+"/"
            if not(os.path.isdir(xl_jo_dir)):
                os.makedirs(os.path.join(xl_jo_dir))
            File_Open_Path(self.xp, xl_jo_dir).jo_writer(data_list,str(name))
        # end of for-loop:name

if __name__ == "__main__":
    
    main_filename_xl = "test_XL.xlsx"

    Xl_to_Json(main_filename_xl).conv()


    # ### Test build

    # main_keys = [
    #     "id",
    #     "name/str",
    #     "skills:list/skill:list/combat",
    #     "skills:list/skill:list/craft",
    #     "skills:list/level/dice:list",
    #     "skills:list/level/add"
    # ]
    # main_vals = [
    #     "id_sample1",
    #     "샘플 이름",
    #     "slash",
    #     "ALL",
    #     [2,6],
    #     4
    # ] 
    
    # # test_conv = Data_Converter(main_keys, main_vals)

    # # print(test_conv.dix_builder())

    # # on openpyxl

    # main_file_px = File_Loader_Opyxl("test_XL.xlsx")

    # main_reading = main_file_px.read_file()
    # main_keys_from_xl = main_file_px.keys_pro(main_reading)
    # main_vals_from_xl = main_file_px.vals_pro(main_reading)

    # names = main_reading.sheetnames

    # xl_conv = Data_Converter()

    # print(main_keys_from_xl["GENERIC"])

    # print(len(main_keys_from_xl["GENERIC"]))

    # print(main_vals_from_xl["GENERIC"][0])
    # print(len(main_vals_from_xl["GENERIC"][0]))

    # print(xl_conv.dix_builder(main_keys_from_xl["GENERIC"],main_vals_from_xl["GENERIC"][0]))

